import os
import re
import json
import datetime
import traceback
import time
import urllib.request
import shutil
import sys
import subprocess
from urllib.error import URLError, HTTPError
from PyQt6.QtCore import QThread, pyqtSignal, QTimer
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                           QLabel, QProgressBar, QFileDialog, QLineEdit, 
                           QTextEdit, QMessageBox, QGroupBox, QApplication)
from PIL import Image, ImageOps
import numpy as np
import imagehash
import cv2

# Define required packages
required_packages = {
    "requests": "requests",
    "PyQt6": "PyQt6",
    "PIL": "Pillow",
    "numpy": "numpy",
    "imagehash": "imagehash",
    "cv2": "opencv-python",
    "pandas": "pandas",
    "openpyxl": "openpyxl"
}

def install_dependencies():
    """Install all required dependencies using pip"""
    missing_packages = []
    
    # Check which packages are missing
    for package, pip_name in required_packages.items():
        try:
            __import__(package)
            print(f"✓ {package} is already installed")
        except ImportError:
            missing_packages.append(pip_name)
            print(f"✗ {package} is missing")
    
    # If there are missing packages, install them
    if missing_packages:
        print("\nInstalling missing dependencies...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing_packages)
            print("\nAll dependencies installed successfully!")
            print("Please restart the application for changes to take effect.")
            return True
        except subprocess.CalledProcessError as e:
            print(f"\nError installing dependencies: {e}")
            print("\nPlease try installing them manually using:")
            print(f"pip install {' '.join(missing_packages)}")
            return False
    else:
        print("\nAll dependencies are already installed!")
        return True

# Check if script is being run with the intention to install dependencies
if len(sys.argv) > 1 and sys.argv[1] == "install_dependencies":
    success = install_dependencies()
    sys.exit(0 if success else 1)

# Check for required external dependencies
missing_packages = []
for package, pip_name in required_packages.items():
    try:
        __import__(package)
    except ImportError:
        missing_packages.append(pip_name)

# If any packages are missing, show error and exit
if missing_packages:
    print("ERROR: Missing required dependencies")
    print("Please install the following packages using pip:")
    print(f"pip install {' '.join(missing_packages)}")
    print("\nOr run this script with the install_dependencies argument:")
    print(f"python {os.path.basename(__file__)} install_dependencies")
    sys.exit(1)

# Import external dependencies after checking
import requests
from PyQt6.QtCore import QThread, pyqtSignal, QTimer
from PyQt6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton, 
                           QLabel, QProgressBar, QFileDialog, QLineEdit, 
                           QTextEdit, QMessageBox, QGroupBox, QApplication)
from PIL import Image, ImageOps
import numpy as np
import imagehash
import cv2

# Settings file path
SETTINGS_FILE = os.path.join(os.path.expanduser("~"), ".marketplace_x_settings.json")

class ImageDownloaderThread(QThread):
    """Thread for downloading images from URLs"""
    progress_update = pyqtSignal(int)
    log_message = pyqtSignal(str)
    download_complete = pyqtSignal()
    
    def __init__(self, urls, output_folder):
        super().__init__()
        self.urls = urls
        self.output_folder = output_folder
        self.cancelled = False
    
    def run(self):
        """Run the download task"""
        # Create download folder
        download_folder = os.path.join(self.output_folder, "download_for_check")
        os.makedirs(download_folder, exist_ok=True)
        
        total_downloaded = 0
        total_errors = 0
        
        for i, url_info in enumerate(self.urls):
            if self.cancelled:
                self.log_message.emit("Download cancelled")
                break
            
            # Get data from URL info
            url = url_info["url"]
            svs = url_info.get("svs", "").zfill(13)  # Ensure 13 digits with leading zeros
            color = url_info.get("color", "")
            position = url_info.get("position", "")
            
            # Skip empty URLs or SVS
            if not url or not svs or svs == "0" * 13:
                continue
            
            # Handle "nan" values that come from pandas missing data
            if color in ["nan", "None", "NaN"] or not color:
                color = ""
            
            try:
                # Construct filename based on available data
                if color and position:
                    filename = f"{svs}_{color}_{position}.jpg"
                elif color:
                    filename = f"{svs}_{color}.jpg"
                else:
                    filename = f"{svs}.jpg"
                
                filepath = os.path.join(download_folder, filename)
                
                # Download image with timeout
                response = requests.get(url, timeout=30)
                
                # Skip only 404 errors
                if response.status_code == 404:
                    self.log_message.emit(f"Image not found (404): {url}")
                    continue
                
                # Save the image for all other responses
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                
                self.log_message.emit(f"Downloaded: {filename}")
                total_downloaded += 1
                
            except Exception as e:
                self.log_message.emit(f"Error downloading {url}: {str(e)}")
                total_errors += 1
            
            # Update progress
            progress = int(100 * (i + 1) / len(self.urls))
            self.progress_update.emit(progress)
        
        self.log_message.emit(f"Download complete. Downloaded: {total_downloaded}, Errors: {total_errors}")
        self.download_complete.emit()

class ImageComparisonThread(QThread):
    """Thread for comparing downloaded images with input images"""
    progress_update = pyqtSignal(int)
    log_message = pyqtSignal(str)
    comparison_complete = pyqtSignal(bool, str, list)
    
    def __init__(self, input_folder, output_folder):
        super().__init__()
        self.input_folder = input_folder
        self.output_folder = output_folder
        self.cancelled = False
        self.comparison_results = []
    
    def compare_histograms(self, img1_path, img2_path):
        """Compare two images using color histograms - simple but effective"""
        try:
            import cv2
            
            # Read images
            img1 = cv2.imread(img1_path)
            img2 = cv2.imread(img2_path)
            
            if img1 is None or img2 is None:
                self.log_message.emit(f"Error reading images")
                return 0
            
            # Calculate histograms for each color channel
            hist_channels = []
            for i in range(3):  # BGR channels
                hist1 = cv2.calcHist([img1], [i], None, [256], [0, 256])
                hist2 = cv2.calcHist([img2], [i], None, [256], [0, 256])
                
                # Normalize histograms
                cv2.normalize(hist1, hist1, 0, 1, cv2.NORM_MINMAX)
                cv2.normalize(hist2, hist2, 0, 1, cv2.NORM_MINMAX)
                
                # Compare histograms - correlation method (higher = more similar)
                similarity = cv2.compareHist(hist1, hist2, cv2.HISTCMP_CORREL)
                hist_channels.append(similarity)
            
            # Average the similarities from all channels
            avg_similarity = sum(hist_channels) / 3 * 100
            
            self.log_message.emit(f"Histogram similarity: {avg_similarity:.2f}%")
            
            return avg_similarity
        
        except Exception as e:
            self.log_message.emit(f"Error in histogram comparison: {str(e)}")
            self.log_message.emit(traceback.format_exc())
            return 0
    
    def compare_image_hashes(self, img1_path, img2_path):
        """Compare two images using perceptual hashing"""
        try:
            # Open images
            img1 = Image.open(img1_path)
            img2 = Image.open(img2_path)
            
            # Calculate perceptual hashes
            hash1 = imagehash.average_hash(img1)
            hash2 = imagehash.average_hash(img2)
            
            # Calculate hash difference
            hash_diff = hash1 - hash2
            
            return hash_diff
        
        except Exception as e:
            self.log_message.emit(f"Error in hash comparison: {str(e)}")
            return None
    
    def compare_images_detailed(self, img1_path, img2_path):
        """
        Strict product image comparison:
        - Resize both images to the same size
        - Use pHash, ORB, and center crop pixel comparison
        - Mark as Duplicate if:
          1. Overall similarity >= 80%, or
          2. Pixel score > 95% with reasonable pHash and ORB scores
        - Otherwise mark as For Retouch
        """
        try:
            import imagehash
            from PIL import Image
            import cv2
            import numpy as np

            # Resize both images to the same size
            size = (512, 512)
            img1_pil = Image.open(img1_path).convert('RGB').resize(size, Image.LANCZOS)
            img2_pil = Image.open(img2_path).convert('RGB').resize(size, Image.LANCZOS)

            # 1. pHash
            hash1 = imagehash.phash(img1_pil)
            hash2 = imagehash.phash(img2_pil)
            hash_diff = hash1 - hash2
            phash_similarity = max(0, 100 - (hash_diff * 10))  # Convert hash diff to similarity percentage

            # 2. ORB feature matching
            img1 = cv2.cvtColor(np.array(img1_pil), cv2.COLOR_RGB2GRAY)
            img2 = cv2.cvtColor(np.array(img2_pil), cv2.COLOR_RGB2GRAY)
            orb = cv2.ORB_create(nfeatures=1000)
            kp1, des1 = orb.detectAndCompute(img1, None)
            kp2, des2 = orb.detectAndCompute(img2, None)
            orb_score = 0
            if des1 is not None and des2 is not None and len(kp1) > 0 and len(kp2) > 0:
                bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
                matches = bf.match(des1, des2)
                good_matches = [m for m in matches if m.distance < 60]
                orb_score = len(good_matches) / max(len(kp1), len(kp2))
            orb_similarity = orb_score * 100  # Convert to percentage

            # 3. Center crop pixel comparison
            arr1 = np.array(img1_pil)
            arr2 = np.array(img2_pil)
            center = (size[0]//2, size[1]//2)
            crop = 128
            crop1 = arr1[center[0]-crop//2:center[0]+crop//2, center[1]-crop//2:center[1]+crop//2]
            crop2 = arr2[center[0]-crop//2:center[0]+crop//2, center[1]-crop//2:center[1]+crop//2]
            pixel_diff = np.mean(np.abs(crop1.astype(float) - crop2.astype(float)))
            pixel_score = 100 - pixel_diff / 2.55

            # Check if both images are almost all white (featureless)
            white1 = np.mean(arr1) > 245
            white2 = np.mean(arr2) > 245
            both_white = white1 and white2

            # Calculate overall similarity score
            if both_white:
                # For white products, only use pHash and ORB
                overall_similarity = (phash_similarity + orb_similarity) / 2
            else:
                # For normal products, use all three methods
                overall_similarity = (phash_similarity + orb_similarity + pixel_score) / 3

            # Decision logic
            is_duplicate = False

            # Case 1: High overall similarity
            if overall_similarity >= 80:
                is_duplicate = True
            # Case 2: High pixel score with reasonable pHash and ORB
            elif pixel_score > 95 and hash_diff <= 18 and orb_score > 0.30:
                is_duplicate = True
            # Case 3: White products need both pHash and ORB to match
            elif both_white and hash_diff <= 5 and orb_score > 0.35:
                is_duplicate = True

            if is_duplicate:
                return 100, f'Duplicate (Similarity: {overall_similarity:.1f}%, pHash={hash_diff}, ORB={orb_score:.2f}, Pixel={pixel_score:.1f})'
            else:
                return int(overall_similarity), f'For Retouch (Similarity: {overall_similarity:.1f}%, pHash={hash_diff}, ORB={orb_score:.2f}, Pixel={pixel_score:.1f})'
        except Exception as e:
            self.log_message.emit(f'Error in robust comparison: {str(e)}')
            return None, f'Error: {str(e)}'
    
    def run(self):
        """Run the comparison task"""
        try:
            # Create folders
            retouch_folder = os.path.join(self.output_folder, "for_retouch")
            potential_duplicates_folder = os.path.join(self.output_folder, "potential_duplicates")
            os.makedirs(retouch_folder, exist_ok=True)
            os.makedirs(potential_duplicates_folder, exist_ok=True)
            
            download_folder = os.path.join(self.output_folder, "download_for_check")
            if not os.path.exists(download_folder):
                self.log_message.emit("Download folder not found! Please download images first.")
                self.comparison_complete.emit(False, "Download folder not found", [])
                return
            
            # Get all input image files
            input_images = []
            for root, _, files in os.walk(self.input_folder):
                for file in files:
                    if file.lower().endswith(('.jpg', '.jpeg', '.png')):
                        input_images.append(os.path.join(root, file))
            
            if not input_images:
                self.log_message.emit("No input images found!")
                self.comparison_complete.emit(False, "No input images found", [])
                return
            
            self.log_message.emit(f"Found {len(input_images)} input images for comparison")
            
            # Get all downloaded files
            downloaded_files = [f for f in os.listdir(download_folder) 
                               if f.lower().endswith(('.jpg', '.jpeg', '.png'))]
            
            self.log_message.emit(f"Found {len(downloaded_files)} downloaded images for comparison")
            
            # Create a dictionary of SVS codes to downloaded files
            svs_to_downloaded = {}
            for filename in downloaded_files:
                # Different pattern matching for downloaded files - more flexible
                svs_match = re.search(r'(\d{13})', filename)
                if svs_match:
                    svs = svs_match.group(1)
                    if svs not in svs_to_downloaded:
                        svs_to_downloaded[svs] = []
                    svs_to_downloaded[svs].append(filename)
            
            self.log_message.emit(f"Indexed {len(svs_to_downloaded)} unique SVS codes in downloaded files")
            
            # Initialize results list
            comparison_results = []
            
            # Process each input image
            self.progress_bar.setValue(0)
            self.status_text.setText("Comparing images...")
            
            # Disable buttons
            self.process_button.setEnabled(False)
            self.download_button.setEnabled(False)
            self.compare_button.setEnabled(False)
            
            total_processed = 0
            for img_path in input_images:
                # Extract filename and SVS
                filename = os.path.basename(img_path)
                svs_match = re.match(r'^(\d{13})', filename)
                
                if not svs_match:
                    self.add_log(f"No SVS found in filename: {filename}")
                    continue
                
                svs = svs_match.group(1)
                
                # Extract color
                color_match = re.search(r'_([A-Za-z0-9]+)', filename)
                color = color_match.group(1) if color_match else ""
                
                # Find matching downloaded files with the same SVS and color
                matching_downloads = []
                for downloaded_file in os.listdir(download_folder):
                    # Check if file starts with SVS
                    if downloaded_file.startswith(svs):
                        # Extract color from downloaded filename
                        downloaded_color_match = re.search(r'_([A-Za-z0-9]+)', downloaded_file)
                        downloaded_color = downloaded_color_match.group(1) if downloaded_color_match else ""
                        
                        # Only add to matches if colors match (or both have no color)
                        if color == downloaded_color:
                            matching_downloads.append(downloaded_file)
                            self.add_log(f"Found matching SVS and color: {downloaded_file} matches color {color}")
                        else:
                            self.add_log(f"SVS matches but color differs - Input: {color}, Downloaded: {downloaded_color}")
                
                # Initialize result
                result = {
                    "Input_Image": filename,
                    "SVS": svs,
                    "Color": color,
                    "Best_Match": "No match",
                    "Status": "For Retouch",
                    "Hash_Difference": None,
                    "Downloaded_Path": "",
                    "Matching_Path": img_path,
                    "Match_Type": "No match"
                }
                
                # If we found matching SVS files, perform detailed comparison
                if matching_downloads:
                    best_similarity = 0
                    best_match = ""
                    best_match_path = ""
                    comparison_reason = ""
                    
                    for downloaded_file in matching_downloads:
                        downloaded_path = os.path.join(download_folder, downloaded_file)
                        similarity_score, reason = self.compare_images_detailed(img_path, downloaded_path)
                        
                        if similarity_score is not None and similarity_score > best_similarity:
                            best_similarity = similarity_score
                            best_match = downloaded_file
                            best_match_path = downloaded_path
                            comparison_reason = reason
                    
                    result["Best_Match"] = best_match
                    result["Similarity_Score"] = best_similarity
                    result["Downloaded_Path"] = best_match_path
                    result["Comparison_Reason"] = comparison_reason
                    
                    # Mark as duplicate only if:
                    # 1. Similarity score is high enough (>70)
                    # 2. No significant differences detected
                    # 3. Reason indicates images are similar
                    if (best_similarity >= 70 and 
                        "differences" not in comparison_reason.lower() and 
                        "not enough" not in comparison_reason.lower()):
                        result["Status"] = "Duplicate"
                        result["Match_Type"] = "Product match"
                        self.add_log(f"Found duplicate: {filename} matches {best_match} (similarity: {best_similarity:.2f}%)")
                    elif best_similarity >= 50:  # Potential duplicate
                        result["Status"] = "Potential Duplicate"
                        result["Match_Type"] = "Potential match"
                        try:
                            dest_path = os.path.join(potential_duplicates_folder, filename)
                            shutil.copy2(img_path, dest_path)
                            self.add_log(f"Copied {filename} to potential duplicates folder (similarity: {best_similarity:.2f}%, reason: {comparison_reason})")
                        except Exception as e:
                            self.add_log(f"Error copying file to potential duplicates folder: {str(e)}")
                    else:
                        # Copy to retouch folder if any differences detected
                        try:
                            dest_path = os.path.join(retouch_folder, filename)
                            shutil.copy2(img_path, dest_path)
                            self.add_log(f"Copied {filename} to retouch folder (similarity: {best_similarity:.2f}%, reason: {comparison_reason})")
                        except Exception as e:
                            self.add_log(f"Error copying file to retouch folder: {str(e)}")
                else:
                    # No matching SVS found, copy to retouch folder
                    try:
                        dest_path = os.path.join(retouch_folder, filename)
                        shutil.copy2(img_path, dest_path)
                        self.add_log(f"Copied {filename} to retouch folder (no matching SVS)")
                    except Exception as e:
                        self.add_log(f"Error copying file to retouch folder: {str(e)}")
                
                comparison_results.append(result)
                
                # Update progress
                total_processed += 1
                progress = int(100 * total_processed / len(input_images))
                self.progress_bar.setValue(progress)
                QApplication.processEvents()
            
            # Generate the comparison Excel file
            try:
                # Try to import pandas
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill
                
                # Create DataFrame
                df = pd.DataFrame(comparison_results)
                
                # Generate output path
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                excel_name = f"Comparison_Results_{timestamp}.xlsx"
                excel_path = os.path.join(self.output_folder, excel_name)
                
                # Save to Excel without styles first
                df.to_excel(excel_path, index=False)
                
                # Now open with openpyxl to add color
                from openpyxl import load_workbook
                wb = load_workbook(excel_path)
                ws = wb.active
                
                # Define color fills
                duplicate_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
                potential_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Yellow
                retouch_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red
                
                # Find status column index
                status_col = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == "Status":
                        status_col = col
                        break
                
                if status_col:
                    # Apply fill colors based on status
                    for row in range(2, ws.max_row + 1):  # Start from row 2 (after header)
                        status = ws.cell(row=row, column=status_col).value
                        if status == "Duplicate":
                            ws.cell(row=row, column=status_col).fill = duplicate_fill
                        elif status == "Potential Duplicate":
                            ws.cell(row=row, column=status_col).fill = potential_fill
                        elif status == "For Retouch":
                            ws.cell(row=row, column=status_col).fill = retouch_fill
                
                # Save workbook with styles
                wb.save(excel_path)
                
                self.add_log(f"Created color-coded comparison results Excel: {excel_path}")
            
            except Exception as e:
                self.add_log(f"Error creating comparison Excel: {str(e)}")
                self.add_log(traceback.format_exc())
                excel_path = ""
            
            # Complete
            retouch_count = sum(1 for r in comparison_results if r["Status"] == "For Retouch")
            duplicate_count = sum(1 for r in comparison_results if r["Status"] == "Duplicate")
            potential_count = sum(1 for r in comparison_results if r["Status"] == "Potential Duplicate")
            
            self.progress_bar.setValue(100)
            self.status_text.setText("Comparison complete")
            
            # Re-enable buttons
            self.process_button.setEnabled(True)
            self.download_button.setEnabled(True)
            self.compare_button.setEnabled(True)
            
            message = f"Comparison complete.\n\n{duplicate_count} duplicates\n{potential_count} potential duplicates\n{retouch_count} need retouch"
            QMessageBox.information(self, "Comparison Complete", message)
            
        except Exception as e:
            self.add_log(f"Comparison error: {str(e)}")
            self.add_log(traceback.format_exc())
            QMessageBox.critical(self, "Error", f"An error occurred:\n{str(e)}")
            
            # Re-enable buttons
            self.process_button.setEnabled(True)
            self.download_button.setEnabled(True)
            self.compare_button.setEnabled(True)

class MarketplaceXWidget(QWidget):
    """MarketplaceX widget for processing images and generating reports"""
    
    def __init__(self, parent=None):
        """Initialize the MarketplaceX widget"""
        super().__init__(parent)
        
        # Initialize storage
        self.input_folder = ""
        self.output_folder = ""
        self.process_start_time = None
        self.excel_path = None
        
        # Initialize thread reference
        self.processor_thread = None
        self.downloader_thread = None
        self.timer = None
        self.comparison_thread = None
        
        # Setup UI
        self.setup_ui()
        
        # Connect signals
        self.connect_signals()
    
    def setup_ui(self):
        """Set up the user interface"""
        # Main layout
        main_layout = QVBoxLayout()
        main_layout.setContentsMargins(10, 10, 10, 10)
        
        # Input folder selection
        input_group = QGroupBox("Input")
        input_layout = QHBoxLayout()
        
        self.input_path_edit = QLineEdit()
        self.input_path_edit.setReadOnly(True)
        self.input_path_edit.setPlaceholderText("Select input folder with images")
        
        self.browse_input_button = QPushButton("Browse")
        
        input_layout.addWidget(self.input_path_edit)
        input_layout.addWidget(self.browse_input_button)
        
        input_group.setLayout(input_layout)
        main_layout.addWidget(input_group)
        
        # Output folder selection
        output_group = QGroupBox("Output")
        output_layout = QHBoxLayout()
        
        self.output_path_edit = QLineEdit()
        self.output_path_edit.setReadOnly(True)
        self.output_path_edit.setPlaceholderText("Select output folder for Excel")
        
        self.browse_output_button = QPushButton("Browse")
        
        output_layout.addWidget(self.output_path_edit)
        output_layout.addWidget(self.browse_output_button)
        
        output_group.setLayout(output_layout)
        main_layout.addWidget(output_group)
        
        # Actions
        actions_group = QGroupBox("Actions")
        actions_layout = QHBoxLayout()
        
        self.process_button = QPushButton("Generate Excel")
        self.download_button = QPushButton("Download Images")
        self.download_button.setEnabled(False)
        self.compare_button = QPushButton("Compare Images")
        self.compare_button.setMinimumHeight(40)
        self.cancel_button = QPushButton("Cancel")
        self.cancel_button.setEnabled(False)
        
        actions_layout.addWidget(self.process_button)
        actions_layout.addWidget(self.download_button)
        actions_layout.addWidget(self.compare_button)
        actions_layout.addWidget(self.cancel_button)
        
        actions_group.setLayout(actions_layout)
        main_layout.addWidget(actions_group)
        
        # Progress
        progress_group = QGroupBox("Progress")
        progress_layout = QVBoxLayout()
        
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        
        self.status_text = QLabel("Ready")
        
        progress_layout.addWidget(self.progress_bar)
        progress_layout.addWidget(self.status_text)
        
        progress_group.setLayout(progress_layout)
        main_layout.addWidget(progress_group)
        
        # Log
        log_group = QGroupBox("Log")
        log_layout = QVBoxLayout()
        
        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)
        
        log_layout.addWidget(self.log_text)
        
        log_group.setLayout(log_layout)
        main_layout.addWidget(log_group)
        
        # Set main layout
        self.setLayout(main_layout)
        
        # Add initial log entry
        self.add_log("Ready")
    
    def connect_signals(self):
        """Connect UI signals to slots"""
        self.browse_input_button.clicked.connect(self.on_browse_input)
        self.browse_output_button.clicked.connect(self.on_browse_output)
        self.process_button.clicked.connect(self.on_process)
        self.download_button.clicked.connect(self.on_download)
        self.compare_button.clicked.connect(self.on_compare)
        self.cancel_button.clicked.connect(self.on_cancel)
    
    def on_browse_input(self):
        """Browse for input folder"""
        folder = QFileDialog.getExistingDirectory(self, "Select Input Folder")
        if folder:
            self.input_folder = folder
            self.input_path_edit.setText(folder)
            self.add_log(f"Input folder: {folder}")
    
    def on_browse_output(self):
        """Browse for output folder"""
        folder = QFileDialog.getExistingDirectory(self, "Select Output Folder")
        if folder:
            self.output_folder = folder
            self.output_path_edit.setText(folder)
            self.add_log(f"Output folder: {folder}")
    
    def on_process(self):
        """Process images and generate Excel file"""
        # Validate inputs
        if not self.input_folder:
            QMessageBox.warning(self, "Input Required", "Please select an input folder first!")
            return
        
        if not self.output_folder:
            QMessageBox.warning(self, "Output Required", "Please select an output folder first!")
            return
        
        # Start processing
        self.add_log("Starting to process images...")
        self.status_text.setText("Processing...")
        self.progress_bar.setValue(10)
        
        # Disable buttons during processing
        self.process_button.setEnabled(False)
        self.download_button.setEnabled(False)
        self.compare_button.setEnabled(False)
        self.cancel_button.setEnabled(True)
        
        try:
            # Get all image files in the input folder
            image_files = []
            for root, _, files in os.walk(self.input_folder):
                for file in files:
                    if file.lower().endswith(('.jpg', '.jpeg', '.png', '.gif')):
                        image_files.append(os.path.join(root, file))
            
            if not image_files:
                self.add_log("No image files found in the input folder!")
                QMessageBox.warning(self, "No Images", "No image files found in the input folder!")
                self.status_text.setText("Ready")
                self.progress_bar.setValue(0)
                self.process_button.setEnabled(True)
                self.cancel_button.setEnabled(False)
                return
            
            self.add_log(f"Found {len(image_files)} image files")
            self.progress_bar.setValue(20)
            
            # Process each image to extract SVS and color
            results = []
            for i, img_path in enumerate(image_files):
                # Extract filename
                filename = os.path.basename(img_path)
                
                # Extract SVS (first 13 digits)
                svs_match = re.match(r'^(\d{13})', filename)
                if svs_match:
                    svs = svs_match.group(1)
                    
                    # Extract color (after underscore)
                    color_match = re.search(r'_([A-Za-z0-9]+)', filename)
                    color = color_match.group(1) if color_match else ""
                    
                    # Create a row with basic info
                    row_data = {
                        "SVS": svs,
                        "Color": color,
                        "Feature_Image_URL": f"http://image.s5a.com/is/image/saks/{svs}?$fullsizejpg$",
                        "Filename": filename
                    }
                    
                    # Generate C-Image URL (only if color is available)
                    if color:
                        row_data["C-Image_URL"] = f"http://image.s5a.com/is/image/saks/{svs}_{color}?$fullsizejpg$"
                    else:
                        row_data["C-Image_URL"] = ""
                    
                    # Generate A1-A20 URLs
                    if color:
                        for angle_num in range(1, 21):
                            angle = f"A{angle_num}"
                            row_data[angle] = f"http://image.s5a.com/is/image/saks/{svs}_{color}_{angle}?$fullsizejpg$"
                    else:
                        # Empty values if no color
                        for angle_num in range(1, 21):
                            angle = f"A{angle_num}"
                            row_data[angle] = ""
                    
                    results.append(row_data)
                    self.add_log(f"Extracted SVS: {svs}, Color: {color} from {filename}")
                else:
                    self.add_log(f"No SVS number found in: {filename}")
                
                # Update progress
                progress = 20 + int(70 * (i + 1) / len(image_files))
                self.progress_bar.setValue(progress)
            
            self.progress_bar.setValue(90)
            
            if not results:
                self.add_log("No SVS numbers found in any images!")
                QMessageBox.warning(self, "No SVS Numbers", "No SVS numbers found in any images!")
                self.status_text.setText("Ready")
                self.progress_bar.setValue(0)
                self.process_button.setEnabled(True)
                self.cancel_button.setEnabled(False)
                return
            
            # Generate Excel file
            self.add_log(f"Generating Excel file with {len(results)} records...")
            
            # Create Excel file
            excel_path = self.create_excel(results)
            if excel_path:
                self.excel_path = excel_path
                self.add_log(f"Excel file created: {excel_path}")
                QMessageBox.information(self, "Success", f"Excel file created successfully:\n{excel_path}")
            else:
                self.add_log("Failed to create Excel file!")
            
            self.progress_bar.setValue(100)
            self.status_text.setText("Complete")
            
            # Enable download button
            self.download_button.setEnabled(True)
            
        except Exception as e:
            self.add_log(f"ERROR: {str(e)}")
            self.add_log(traceback.format_exc())
            QMessageBox.critical(self, "Error", f"An error occurred:\n{str(e)}")
            self.progress_bar.setValue(0)
            self.status_text.setText("Error")
        
        # Re-enable buttons
        self.process_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
    
    def on_download(self):
        """Download images for URLs from the Excel file"""
        if not self.output_folder:
            QMessageBox.warning(self, "Output Required", "Please select an output folder first!")
            return

        if not self.excel_path:
            QMessageBox.warning(self, "No Excel", 
                              "No Excel file available for download!\n\n" +
                              "Please process an Excel file first.")
            return
        
        # Create download folder
        download_folder = os.path.join(self.output_folder, "download_for_check")
        os.makedirs(download_folder, exist_ok=True)
        
        try:
            # Read URLs from Excel file
            import pandas as pd
            df = pd.read_excel(self.excel_path)
            
            # Create a list of all URLs to download
            download_urls = []
            
            # Check which columns exist in the dataframe
            url_columns = [col for col in df.columns if 'URL' in col or 'url' in col or col.startswith('A')]
            
            # Process each row in the Excel file
            for _, row in df.iterrows():
                # Ensure SVS is a 13-digit string with leading zeros
                svs = str(row.get("SVS", "")).zfill(13)
                color = str(row.get("Color", ""))
                
                # Skip empty SVS
                if not svs or svs == "0" * 13:
                    continue
                
                # Add Feature Image URL if it exists
                if "Feature_Image_URL" in df.columns and pd.notna(row["Feature_Image_URL"]):
                    download_urls.append({
                        "url": row["Feature_Image_URL"],
                        "svs": svs,  # Now properly formatted
                        "color": color,
                        "position": ""
                    })
                
                # Add C-Image URL if it exists
                if "C-Image_URL" in df.columns and pd.notna(row["C-Image_URL"]):
                    download_urls.append({
                        "url": row["C-Image_URL"],
                        "svs": svs,  # Now properly formatted
                        "color": color,
                        "position": ""
                    })
                
                # Add angle URLs (A1-A20)
                for angle_num in range(1, 21):
                    angle = f"A{angle_num}"
                    if angle in df.columns and pd.notna(row[angle]):
                        download_urls.append({
                            "url": row[angle],
                            "svs": svs,  # Now properly formatted
                            "color": color,
                            "position": angle
                        })
            
            if not download_urls:
                QMessageBox.warning(self, "No URLs", "No valid URLs found for download!")
                return
            
            self.add_log(f"Found {len(download_urls)} URLs to download")
            
            # Start download thread
            self.downloader_thread = ImageDownloaderThread(download_urls, self.output_folder)
            
            # Connect signals
            self.downloader_thread.progress_update.connect(self.progress_bar.setValue)
            self.downloader_thread.log_message.connect(self.add_log)
            self.downloader_thread.download_complete.connect(self.on_download_complete)
            
            # Update UI
            self.progress_bar.setValue(0)
            self.status_text.setText("Downloading images...")
            
            # Disable buttons
            self.process_button.setEnabled(False)
            self.download_button.setEnabled(False)
            self.compare_button.setEnabled(False)
            self.cancel_button.setEnabled(True)
            
            # Start thread
            self.downloader_thread.start()
            
        except Exception as e:
            self.add_log(f"Error setting up download: {str(e)}")
            self.add_log(traceback.format_exc())
            QMessageBox.critical(self, "Error", f"An error occurred:\n{str(e)}")
    
    def on_download_complete(self):
        """Handle completion of image download"""
        # Update UI
        self.process_button.setEnabled(True)
        self.download_button.setEnabled(True)
        self.compare_button.setEnabled(True)
        self.cancel_button.setEnabled(False)
        
        self.progress_bar.setValue(100)
        self.status_text.setText("Download Complete")
        
        # Show success message
        QMessageBox.information(
            self, "Download Complete", 
            f"All images downloaded successfully!"
        )
    
    def on_compare(self):
        """Compare downloaded images with input images using perceptual hashing"""
        # Validate inputs
        if not self.input_folder:
            QMessageBox.warning(self, "Input Required", "Please select an input folder first!")
            return
        
        if not self.output_folder:
            QMessageBox.warning(self, "Output Required", "Please select an output folder first!")
            return
        
        # Check if download folder exists
        download_folder = os.path.join(self.output_folder, "download_for_check")
        if not os.path.exists(download_folder):
            QMessageBox.warning(self, "Download Required", 
                              "Please download images first!\n\n" +
                              "The download folder does not exist.")
            return
        
        try:
            # Create retouch folder
            retouch_folder = os.path.join(self.output_folder, "for_retouch")
            os.makedirs(retouch_folder, exist_ok=True)
            
            self.add_log(f"Created retouch folder: {retouch_folder}")
            
            # Get all input image files
            input_images = []
            for root, _, files in os.walk(self.input_folder):
                for file in files:
                    if file.lower().endswith(('.jpg', '.jpeg', '.png')):
                        input_images.append(os.path.join(root, file))
            
            if not input_images:
                self.add_log("No input images found!")
                QMessageBox.warning(self, "No Images", "No input images found!")
                return
            
            self.add_log(f"Found {len(input_images)} input images for comparison")
            
            # Initialize results list
            comparison_results = []
            
            # Process each input image
            self.progress_bar.setValue(0)
            self.status_text.setText("Comparing images...")
            
            # Disable buttons
            self.process_button.setEnabled(False)
            self.download_button.setEnabled(False)
            self.compare_button.setEnabled(False)
            
            # Set hash difference threshold
            HASH_THRESHOLD = 10  # Lower number means more similar
            
            total_processed = 0
            for img_path in input_images:
                # Extract filename and SVS
                filename = os.path.basename(img_path)
                svs_match = re.match(r'^(\d{13})', filename)
                
                if not svs_match:
                    self.add_log(f"No SVS found in filename: {filename}")
                    continue
                
                svs = svs_match.group(1)
                
                # Extract color
                color_match = re.search(r'_([A-Za-z0-9]+)', filename)
                color = color_match.group(1) if color_match else ""
                
                # Find matching downloaded files with the same SVS and color
                matching_downloads = []
                for downloaded_file in os.listdir(download_folder):
                    # Check if file starts with SVS
                    if downloaded_file.startswith(svs):
                        # Extract color from downloaded filename
                        downloaded_color_match = re.search(r'_([A-Za-z0-9]+)', downloaded_file)
                        downloaded_color = downloaded_color_match.group(1) if downloaded_color_match else ""
                        
                        # Only add to matches if colors match (or both have no color)
                        if color == downloaded_color:
                            matching_downloads.append(downloaded_file)
                            self.add_log(f"Found matching SVS and color: {downloaded_file} matches color {color}")
                        else:
                            self.add_log(f"SVS matches but color differs - Input: {color}, Downloaded: {downloaded_color}")
                
                # Initialize result
                result = {
                    "Input_Image": filename,
                    "SVS": svs,
                    "Color": color,
                    "Best_Match": "No match",
                    "Status": "For Retouch",
                    "Hash_Difference": None,
                    "Downloaded_Path": "",
                    "Matching_Path": img_path,
                    "Match_Type": "No match"
                }
                
                # If we found matching SVS files, perform detailed comparison
                if matching_downloads:
                    best_similarity = 0
                    best_match = ""
                    best_match_path = ""
                    comparison_reason = ""
                    
                    for downloaded_file in matching_downloads:
                        downloaded_path = os.path.join(download_folder, downloaded_file)
                        similarity_score, reason = self.compare_images_detailed(img_path, downloaded_path)
                        
                        if similarity_score is not None and similarity_score > best_similarity:
                            best_similarity = similarity_score
                            best_match = downloaded_file
                            best_match_path = downloaded_path
                            comparison_reason = reason
                    
                    result["Best_Match"] = best_match
                    result["Similarity_Score"] = best_similarity
                    result["Downloaded_Path"] = best_match_path
                    result["Comparison_Reason"] = comparison_reason
                    
                    # Mark as duplicate only if:
                    # 1. Similarity score is high enough (>70)
                    # 2. No significant differences detected
                    # 3. Reason indicates images are similar
                    if (best_similarity >= 70 and 
                        "differences" not in comparison_reason.lower() and 
                        "not enough" not in comparison_reason.lower()):
                        result["Status"] = "Duplicate"
                        result["Match_Type"] = "Product match"
                        self.add_log(f"Found duplicate: {filename} matches {best_match} (similarity: {best_similarity:.2f}%)")
                    elif best_similarity >= 50:  # Potential duplicate
                        result["Status"] = "Potential Duplicate"
                        result["Match_Type"] = "Potential match"
                        try:
                            dest_path = os.path.join(retouch_folder, filename)
                            shutil.copy2(img_path, dest_path)
                            self.add_log(f"Copied {filename} to retouch folder (similarity: {best_similarity:.2f}%, reason: {comparison_reason})")
                        except Exception as e:
                            self.add_log(f"Error copying file to retouch folder: {str(e)}")
                    else:
                        # Copy to retouch folder if any differences detected
                        try:
                            dest_path = os.path.join(retouch_folder, filename)
                            shutil.copy2(img_path, dest_path)
                            self.add_log(f"Copied {filename} to retouch folder (similarity: {best_similarity:.2f}%, reason: {comparison_reason})")
                        except Exception as e:
                            self.add_log(f"Error copying file to retouch folder: {str(e)}")
                else:
                    # No matching SVS found, copy to retouch folder
                    try:
                        dest_path = os.path.join(retouch_folder, filename)
                        shutil.copy2(img_path, dest_path)
                        self.add_log(f"Copied {filename} to retouch folder (no matching SVS)")
                    except Exception as e:
                        self.add_log(f"Error copying file to retouch folder: {str(e)}")
                
                comparison_results.append(result)
                
                # Update progress
                total_processed += 1
                progress = int(100 * total_processed / len(input_images))
                self.progress_bar.setValue(progress)
                QApplication.processEvents()
            
            # Generate the comparison Excel file
            try:
                # Try to import pandas
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill
                
                # Create DataFrame
                df = pd.DataFrame(comparison_results)
                
                # Generate output path
                timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                excel_name = f"Comparison_Results_{timestamp}.xlsx"
                excel_path = os.path.join(self.output_folder, excel_name)
                
                # Save to Excel without styles first
                df.to_excel(excel_path, index=False)
                
                # Now open with openpyxl to add color
                from openpyxl import load_workbook
                wb = load_workbook(excel_path)
                ws = wb.active
                
                # Define color fills
                duplicate_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Green
                potential_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Yellow
                retouch_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red
                
                # Find status column index
                status_col = None
                for col in range(1, ws.max_column + 1):
                    if ws.cell(row=1, column=col).value == "Status":
                        status_col = col
                        break
                
                if status_col:
                    # Apply fill colors based on status
                    for row in range(2, ws.max_row + 1):  # Start from row 2 (after header)
                        status = ws.cell(row=row, column=status_col).value
                        if status == "Duplicate":
                            ws.cell(row=row, column=status_col).fill = duplicate_fill
                        elif status == "Potential Duplicate":
                            ws.cell(row=row, column=status_col).fill = potential_fill
                        elif status == "For Retouch":
                            ws.cell(row=row, column=status_col).fill = retouch_fill
                
                # Save workbook with styles
                wb.save(excel_path)
                
                self.add_log(f"Created color-coded comparison results Excel: {excel_path}")
            
            except Exception as e:
                self.add_log(f"Error creating comparison Excel: {str(e)}")
                self.add_log(traceback.format_exc())
                excel_path = ""
            
            # Complete
            retouch_count = sum(1 for r in comparison_results if r["Status"] == "For Retouch")
            duplicate_count = sum(1 for r in comparison_results if r["Status"] == "Duplicate")
            potential_count = sum(1 for r in comparison_results if r["Status"] == "Potential Duplicate")
            
            self.progress_bar.setValue(100)
            self.status_text.setText("Comparison complete")
            
            # Re-enable buttons
            self.process_button.setEnabled(True)
            self.download_button.setEnabled(True)
            self.compare_button.setEnabled(True)
            
            message = f"Comparison complete.\n\n{duplicate_count} duplicates\n{potential_count} potential duplicates\n{retouch_count} need retouch"
            QMessageBox.information(self, "Comparison Complete", message)
            
        except Exception as e:
            self.add_log(f"Comparison error: {str(e)}")
            self.add_log(traceback.format_exc())
            QMessageBox.critical(self, "Error", f"An error occurred:\n{str(e)}")
            
            # Re-enable buttons
            self.process_button.setEnabled(True)
            self.download_button.setEnabled(True)
            self.compare_button.setEnabled(True)
    
    def on_cancel(self):
        """Cancel current operation"""
        if hasattr(self, 'comparison_thread') and self.comparison_thread and self.comparison_thread.isRunning():
            self.comparison_thread.cancelled = True
            self.add_log("Cancelling comparison...")
            self.status_text.setText("Cancelling...")
            self.cancel_button.setEnabled(False)
        if self.downloader_thread and self.downloader_thread.isRunning():
            self.downloader_thread.cancelled = True
            self.add_log("Cancelling download...")
            self.status_text.setText("Cancelling...")
            self.cancel_button.setEnabled(False)
    
    def create_excel(self, data):
        """Create Excel file with SVS and Color data"""
        try:
            # Try to import pandas
            try:
                import pandas as pd
            except ImportError:
                self.add_log("pandas is not installed. Please install it with: pip install pandas openpyxl")
                QMessageBox.warning(self, "Missing Dependency", 
                                  "pandas is not installed.\n\nPlease install it with:\npip install pandas openpyxl")
                return None
            
            # Create DataFrame
            df = pd.DataFrame(data)
            
            # Generate output path
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            excel_name = f"SVS_Report_{timestamp}.xlsx"
            excel_path = os.path.join(self.output_folder, excel_name)
            
            # Write to Excel
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)
            
            return excel_path
        
        except Exception as e:
            self.add_log(f"Error creating Excel: {str(e)}")
            self.add_log(traceback.format_exc())
            return None
    
    def add_log(self, message):
        """Add message to log"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}"
        self.log_text.append(log_message)
        
        # Scroll to bottom
        scrollbar = self.log_text.verticalScrollBar()
        scrollbar.setValue(scrollbar.maximum())
    
    def update_progress(self, value):
        """Update progress bar"""
        self.progress_bar.setValue(value)
    
    def compare_images_detailed(self, img1_path, img2_path):
        """
        Strict product image comparison:
        - Resize both images to the same size
        - Use pHash, ORB, and center crop pixel comparison
        - Mark as Duplicate if:
          1. Overall similarity >= 80%, or
          2. Pixel score > 95% with reasonable pHash and ORB scores
        - Otherwise mark as For Retouch
        """
        try:
            import imagehash
            from PIL import Image
            import cv2
            import numpy as np

            # Resize both images to the same size
            size = (512, 512)
            img1_pil = Image.open(img1_path).convert('RGB').resize(size, Image.LANCZOS)
            img2_pil = Image.open(img2_path).convert('RGB').resize(size, Image.LANCZOS)

            # 1. pHash
            hash1 = imagehash.phash(img1_pil)
            hash2 = imagehash.phash(img2_pil)
            hash_diff = hash1 - hash2
            phash_similarity = max(0, 100 - (hash_diff * 10))  # Convert hash diff to similarity percentage

            # 2. ORB feature matching
            img1 = cv2.cvtColor(np.array(img1_pil), cv2.COLOR_RGB2GRAY)
            img2 = cv2.cvtColor(np.array(img2_pil), cv2.COLOR_RGB2GRAY)
            orb = cv2.ORB_create(nfeatures=1000)
            kp1, des1 = orb.detectAndCompute(img1, None)
            kp2, des2 = orb.detectAndCompute(img2, None)
            orb_score = 0
            if des1 is not None and des2 is not None and len(kp1) > 0 and len(kp2) > 0:
                bf = cv2.BFMatcher(cv2.NORM_HAMMING, crossCheck=True)
                matches = bf.match(des1, des2)
                good_matches = [m for m in matches if m.distance < 60]
                orb_score = len(good_matches) / max(len(kp1), len(kp2))
            orb_similarity = orb_score * 100  # Convert to percentage

            # 3. Center crop pixel comparison
            arr1 = np.array(img1_pil)
            arr2 = np.array(img2_pil)
            center = (size[0]//2, size[1]//2)
            crop = 128
            crop1 = arr1[center[0]-crop//2:center[0]+crop//2, center[1]-crop//2:center[1]+crop//2]
            crop2 = arr2[center[0]-crop//2:center[0]+crop//2, center[1]-crop//2:center[1]+crop//2]
            pixel_diff = np.mean(np.abs(crop1.astype(float) - crop2.astype(float)))
            pixel_score = 100 - pixel_diff / 2.55

            # Check if both images are almost all white (featureless)
            white1 = np.mean(arr1) > 245
            white2 = np.mean(arr2) > 245
            both_white = white1 and white2

            # Calculate overall similarity score
            if both_white:
                # For white products, only use pHash and ORB
                overall_similarity = (phash_similarity + orb_similarity) / 2
            else:
                # For normal products, use all three methods
                overall_similarity = (phash_similarity + orb_similarity + pixel_score) / 3

            # Decision logic
            is_duplicate = False

            # Case 1: High overall similarity
            if overall_similarity >= 80:
                is_duplicate = True
            # Case 2: High pixel score with reasonable pHash and ORB
            elif pixel_score > 95 and hash_diff <= 18 and orb_score > 0.30:
                is_duplicate = True
            # Case 3: White products need both pHash and ORB to match
            elif both_white and hash_diff <= 5 and orb_score > 0.35:
                is_duplicate = True

            if is_duplicate:
                return 100, f'Duplicate (Similarity: {overall_similarity:.1f}%, pHash={hash_diff}, ORB={orb_score:.2f}, Pixel={pixel_score:.1f})'
            else:
                return int(overall_similarity), f'For Retouch (Similarity: {overall_similarity:.1f}%, pHash={hash_diff}, ORB={orb_score:.2f}, Pixel={pixel_score:.1f})'
        except Exception as e:
            self.add_log(f'Error in robust comparison: {str(e)}')
            return None, f'Error: {str(e)}'
